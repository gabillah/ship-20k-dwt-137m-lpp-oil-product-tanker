"""
excel_to_latex.py
=================
Membaca nilai dari calculation_p1.xlsx lalu menulis calc_values.tex
berisi \newcommand yang siap di-\input{} oleh report.tex.

Penggunaan::

    python excel_to_latex.py                       # pakai path default
    python excel_to_latex.py myfile.xlsx           # custom Excel file
    python excel_to_latex.py myfile.xlsx out.tex   # custom keduanya

Dependensi:
    pip install pandas openpyxl
"""

import sys
import os
import math
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook


# ─────────────────────────────────────────────────────────────────
# Konfigurasi path (bisa di-override via argumen CLI)
# ─────────────────────────────────────────────────────────────────
EXCEL_FILE = sys.argv[1] if len(sys.argv) > 1 else "calculation_p1.xlsx"
TEX_OUTPUT = sys.argv[2] if len(sys.argv) > 2 else "calc_values.tex"

# Pastikan output ditulis di folder yang sama dengan Excel
BASE_DIR   = os.path.dirname(os.path.abspath(EXCEL_FILE))
TEX_OUTPUT = os.path.join(BASE_DIR, os.path.basename(TEX_OUTPUT))


# ─────────────────────────────────────────────────────────────────
# Helper: format angka untuk LaTeX
# ─────────────────────────────────────────────────────────────────
def fmt(value: float, decimals: int = 4) -> str:
    """Format float ke string desimal dengan titik sebagai separator."""
    return f"{value:.{decimals}f}"


def fmt_sci(value: float, decimals: int = 3) -> str:
    """Scientific notation untuk LaTeX: 3.000 \\times 10^{5}"""
    if value == 0:
        return "0"
    exp = int(math.floor(math.log10(abs(value))))
    mantissa = value / (10 ** exp)
    return rf"{mantissa:.{decimals}f} \times 10^{{{exp}}}"


def read_cell(ws, cell_address: str) -> float:
    """Baca nilai numerik dari sel (formula dievaluasi jika file di-recalc)."""
    val = ws[cell_address].value
    if val is None:
        return 0.0
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0.0


# ─────────────────────────────────────────────────────────────────
# Baca Excel
# ─────────────────────────────────────────────────────────────────
print(f"Membaca: {EXCEL_FILE}")

# data_only=True -> ambil nilai hasil kalkulasi, bukan string formula
wb = load_workbook(EXCEL_FILE, data_only=True)

ws_inputs  = wb["Inputs"]
ws_results = wb["Results"]

# ── Inputs (sel biru — nilai hardcoded, selalu bisa dibaca) ──────
a          = read_cell(ws_inputs, "B3")   # Operand a
b          = read_cell(ws_inputs, "B4")   # Operand b
f_friction = read_cell(ws_inputs, "B6")   # Friction factor
L_pipe     = read_cell(ws_inputs, "B7")   # Pipe length [m]
D_pipe     = read_cell(ws_inputs, "B8")   # Pipe diameter [m]
V_flow     = read_cell(ws_inputs, "B9")   # Flow velocity [m/s]
g_gravity  = read_cell(ws_inputs, "B10")  # Gravity [m/s²]

# ── Results (formula Excel — tersedia jika file pernah dibuka/disave di Excel) ──
# Fallback: hitung ulang di Python jika nilai None (file belum pernah dibuka Excel)
c_excel  = read_cell(ws_results, "B2")
Re_excel = read_cell(ws_results, "B3")
hf_excel = read_cell(ws_results, "B4")

# Python fallback calculation
c_py  = a + b
Re_py = (V_flow * D_pipe) / 1e-6
hf_py = f_friction * (L_pipe / D_pipe) * (V_flow**2 / (2 * g_gravity))

# Pakai nilai Excel jika tersedia (> 0), kalau tidak pakai Python
c  = c_excel  if c_excel  != 0.0 else c_py
Re = Re_excel if Re_excel != 0.0 else Re_py
hf = hf_excel if hf_excel != 0.0 else hf_py

print(f"  a={a}, b={b}, c={c}")
print(f"  f={f_friction}, L={L_pipe}, D={D_pipe}, V={V_flow}")
print(f"  Re={Re:.0f}, hf={hf:.4f} m")


# ─────────────────────────────────────────────────────────────────
# Tulis calc_values.tex
# ─────────────────────────────────────────────────────────────────
timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

lines = [
    r"% ============================================================",
    rf"% AUTO-GENERATED oleh excel_to_latex.py — {timestamp}",
    rf"% Sumber Excel : {os.path.basename(EXCEL_FILE)}",
    r"% JANGAN EDIT MANUAL — jalankan script untuk regenerasi.",
    r"% ============================================================",
    "",
    r"% ── Section 1: Simple Addition ─────────────────────────────",
    rf"\newcommand{{\ValA}}{{{fmt(a, 2)}}}",
    rf"\newcommand{{\ValB}}{{{fmt(b, 2)}}}",
    rf"\newcommand{{\ValC}}{{{fmt(c, 2)}}}",
    "",
    r"% ── Section 2: Darcy-Weisbach ──────────────────────────────",
    rf"\newcommand{{\ValFriction}}{{{fmt(f_friction, 3)}}}",
    rf"\newcommand{{\ValLpipe}}{{{fmt(L_pipe, 1)}}}",
    rf"\newcommand{{\ValDpipe}}{{{fmt(D_pipe, 2)}}}",
    rf"\newcommand{{\ValVflow}}{{{fmt(V_flow, 1)}}}",
    rf"\newcommand{{\ValRe}}{{{Re:.0f}}}",
    rf"\newcommand{{\ValHf}}{{{fmt(hf, 4)}}}",
    "",
    r"% ── Scientific notation ─────────────────────────────────────",
    rf"\newcommand{{\ValReSci}}{{{fmt_sci(Re)}}}",
]

snippet = "\n".join(lines)

with open(TEX_OUTPUT, "w", encoding="utf-8") as f:
    f.write(snippet)

print(f"\nDitulis: {TEX_OUTPUT}")
print("─" * 60)
print(snippet)
print("─" * 60)
print("\n✓ Selesai. Compile report.tex dengan XeLaTeX.")
